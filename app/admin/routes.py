from flask import render_template, request, flash, redirect, url_for, current_app, send_file, abort
from flask_login import login_required, current_user
from sqlalchemy import func, distinct
from . import admin
from .. import db, bcrypt
from ..models import (Product, Category, ContactMessage, StaffUser, Order, OrderItem, Customer, 
                     ProductImage, Post, PageContent, Banner, Milestone, Newsletter, NewsletterSubscriber)
from ..forms import (CategoryForm, ProductForm, DeleteForm, StaffUserEditForm, 
                   ContactMessageEditForm, ReplyForm, CustomerEditForm, StaffRegistrationForm, PostForm, PageContentForm, BannerForm, MilestoneForm, NewsletterCreationForm, SendForm)
from ..utils.image_helpers import save_image, allowed_file, delete_image_from_cloudinary
from openpyxl import Workbook
from io import BytesIO
from functools import wraps
from werkzeug.datastructures import FileStorage
from datetime import date, timedelta, datetime, timezone
import pytz
import os
from flask_mailman import EmailMessage

def merge_query_args(args, new_args):
    args = args.copy()
    for key, value in new_args.items():
        args[key] = value
    return args

# Decorators
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not isinstance(current_user, StaffUser) or current_user.role != 'admin':
            flash("Vous n'avez pas les permissions nécessaires pour accéder à cette page.", 'danger')
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated_function

def staff_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if current_app.config.get('LOGIN_DISABLED'):
            return f(*args, **kwargs)
        if not current_user.is_authenticated or not isinstance(current_user, StaffUser):
            flash("Vous devez être connecté en tant que membre du personnel pour accéder à cette page.", 'danger')
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated_function

def customer_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not isinstance(current_user, Customer):
            flash("Cette page est réservée aux clients.", "danger")
            return redirect(url_for('main.index'))
        return f(*args, **kwargs)
    return decorated_function

@admin.route('/dashboard')
@staff_required
def admin_dashboard():
    # --- Date Range Filter ---
    end_date = datetime.now(timezone.utc).date()
    start_date = end_date - timedelta(days=29) # Default to last 30 days

    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except ValueError:
            flash('Format de date de début invalide. Utilisez AAAA-MM-JJ.', 'danger')
    
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            flash('Format de date de fin invalide. Utilisez AAAA-MM-JJ.', 'danger')

    # --- Stats Calculation ---
    orders_in_range = db.session.execute(db.select(Order).filter(Order.date_ordered.between(start_date, end_date + timedelta(days=1)))).scalars()
    
    total_revenue = db.session.query(db.func.sum(Order.total_price)).filter(Order.date_ordered.between(start_date, end_date + timedelta(days=1))).scalar() or 0
    total_orders = db.session.execute(db.select(func.count()).select_from(Order).filter(Order.date_ordered.between(start_date, end_date + timedelta(days=1)))).scalar_one()
    
    total_customers = db.session.query(func.count(Customer.id)).scalar()
    total_products = db.session.query(func.count(Product.id)).scalar()

    today = datetime.now(timezone.utc).date()
    revenue_today = db.session.query(db.func.sum(Order.total_price)).filter(db.func.date(Order.date_ordered) == today).scalar() or 0
    orders_today = db.session.query(func.count(Order.id)).filter(db.func.date(Order.date_ordered) == today).scalar()

    low_stock_products = db.session.execute(db.select(Product).filter(Product.stock <= Product.min_stock_threshold)).scalars().all()
    latest_orders = db.session.execute(db.select(Order).order_by(Order.date_ordered.desc()).limit(5)).scalars().all()

    top_selling_products = db.session.query(
        Product.name,
        db.func.sum(OrderItem.quantity).label('total_quantity')
    ).join(OrderItem).join(Order).filter(Order.date_ordered.between(start_date, end_date + timedelta(days=1))).group_by(Product.name).order_by(db.desc('total_quantity')).limit(5).all()

    # --- Sales Chart ---
    delta = end_date - start_date
    sales_data = {}
    for i in range(delta.days + 1):
        day = start_date + timedelta(days=i)
        sales_data[day.strftime('%Y-%m-%d')] = 0

    sales_in_range = db.session.query(
        db.func.date(Order.date_ordered),
        db.func.sum(Order.total_price)
    ).filter(Order.date_ordered.between(start_date, end_date + timedelta(days=1))).group_by(db.func.date(Order.date_ordered)).all()

    for day, total in sales_in_range:
        if day:
            sales_data[day] = total
    
    chart_labels = list(sales_data.keys())
    chart_values = list(sales_data.values())

    # --- New Customers Chart ---
    new_customers_data = {}
    for i in range(delta.days + 1):
        day = start_date + timedelta(days=i)
        new_customers_data[day.strftime('%Y-%m-%d')] = 0

    new_customers_in_range = db.session.query(
        db.func.date(Customer.date_registered),
        db.func.count(Customer.id)
    ).filter(Customer.date_registered.between(start_date, end_date + timedelta(days=1))).group_by(db.func.date(Customer.date_registered)).all()

    for day, count in new_customers_in_range:
        if day:
            new_customers_data[day] = count
    
    new_customers_chart_labels = list(new_customers_data.keys())
    new_customers_chart_values = list(new_customers_data.values())

    # --- Page Visit Stats ---
    from ..models import PageVisit 
    sessions_in_range = db.session.query(func.count(distinct(PageVisit.session_id))).filter(PageVisit.timestamp.between(start_date, end_date + timedelta(days=1))).scalar() or 0
    page_views_in_range = db.session.query(func.count(PageVisit.id)).filter(PageVisit.timestamp.between(start_date, end_date + timedelta(days=1))).scalar() or 0

    return render_template('admin_dashboard.html', 
                           total_revenue=total_revenue, 
                           total_orders=total_orders,
                           total_customers=total_customers,
                           total_products=total_products,
                           revenue_today=revenue_today,
                           orders_today=orders_today,
                           low_stock_products=low_stock_products,
                           latest_orders=latest_orders,
                           top_selling_products=top_selling_products,
                           chart_labels=chart_labels,
                           chart_values=chart_values,
                           new_customers_chart_labels=new_customers_chart_labels,
                           new_customers_chart_values=new_customers_chart_values,
                           sessions_today=sessions_in_range,
                           page_views_today=page_views_in_range,
                           filters={
                               'start_date': start_date.strftime('%Y-%m-%d'),
                               'end_date': end_date.strftime('%Y-%m-%d')
                           })

@admin.route('/posts')
@staff_required
def admin_posts():
    posts = db.session.execute(db.select(Post).order_by(Post.created_at.desc())).scalars().all()
    delete_form = DeleteForm()
    
    # Convertir les dates UTC en heure locale de Paris et les formater
    paris_tz = pytz.timezone('Europe/Paris')
    for post in posts:
        if post.created_at:
            utc_dt = pytz.utc.localize(post.created_at)
            paris_dt = utc_dt.astimezone(paris_tz)
            post.created_at_formatted = paris_dt.strftime('%d/%m/%Y à %H:%M')
        else:
            post.created_at_formatted = "N/A"

    return render_template('admin/posts.html', posts=posts, delete_form=delete_form)

@admin.route('/post/add', methods=['GET', 'POST'])
@staff_required
def add_post():
    form = PostForm()
    if form.validate_on_submit():
        new_post = Post(
            title=form.title.data,
            description=form.description.data,
            video_url=form.video_url.data,
            author_id=current_user.id,
            created_at=datetime.now(timezone.utc)
        )
        if form.cover_image.data:
            try:
                filename = save_image(form.cover_image.data, current_app.config['UPLOAD_FOLDER'])
                new_post.cover_image = filename
            except Exception as e:
                flash(f"Une erreur est survenue lors de l'enregistrement de l'image : {e}", 'danger')
                return render_template('admin/add_post.html', form=form)
        
        db.session.add(new_post)
        db.session.commit()
        flash('La réalisation a été ajoutée avec succès !', 'success')
        return redirect(url_for('admin.admin_posts'))
    return render_template('admin/add_post.html', form=form, title="Ajouter une réalisation")

@admin.route('/post/edit/<int:post_id>', methods=['GET', 'POST'])
@staff_required
def edit_post(post_id):
    post = db.session.get(Post, post_id) or abort(404)
    form = PostForm(obj=post)
    if form.validate_on_submit():
        post.title = form.title.data
        post.description = form.description.data
        post.video_url = form.video_url.data
        
        if form.cover_image.data and isinstance(form.cover_image.data, FileStorage):
            try:
                if post.cover_image and 'cloudinary' in post.cover_image:
                    delete_image_from_cloudinary(post.cover_image)
                
                filename = save_image(form.cover_image.data, current_app.config['UPLOAD_FOLDER'])
                post.cover_image = filename
            except Exception as e:
                flash(f"Une erreur est survenue lors de l'enregistrement de la nouvelle image : {e}", 'danger')
                return render_template('admin/add_post.html', form=form, title="Modifier la réalisation", post=post)

        db.session.commit()
        flash('La réalisation a été mise à jour avec succès !', 'success')
        return redirect(url_for('admin.admin_posts'))

    return render_template('admin/add_post.html', form=form, title="Modifier la réalisation", post=post)

@admin.route('/categories')
@staff_required
def admin_categories():
    delete_form = DeleteForm()
    categories = db.session.execute(db.select(Category)).scalars().all()
    return render_template('admin_categories.html', categories=categories, delete_form=delete_form)

@admin.route('/category/add', methods=['GET', 'POST'])
@staff_required
def add_category():
    form = CategoryForm()
    if form.validate_on_submit():
        name = form.name.data
        if db.session.execute(db.select(Category).filter_by(name=name)).scalar_one_or_none():
            flash("Cette catégorie existe déjà.", 'danger')
        else:
            new_category = Category(name=name)
            db.session.add(new_category)
            db.session.commit()
            flash("La catégorie a été ajoutée avec succès.", 'success')
            return redirect(url_for('admin.admin_categories'))
    return render_template('add_category.html', form=form)

@admin.route('/category/edit/<int:category_id>', methods=['GET', 'POST'])
@staff_required
def edit_category(category_id):
    category = db.session.get(Category, category_id)
    if not category:
        flash("Catégorie introuvable.", "danger")
        return redirect(url_for('admin.admin_categories'))
    form = CategoryForm(obj=category)
    if form.validate_on_submit():
        category.name = form.name.data
        db.session.commit()
        flash("La catégorie a été mise à jour avec succès.", "success")
        return redirect(url_for('admin.admin_categories'))
    return render_template('add_category.html', form=form, category=category)

@admin.route('/category/delete/<int:category_id>', methods=['POST'])
@admin_required
def delete_category(category_id):
    category = db.session.get(Category, category_id)
    if category:
        db.session.delete(category)
        db.session.commit()
        flash("La catégorie a été supprimée avec succès.", "success")
    else:
        flash("Catégorie introuvable.", "danger")
    return redirect(url_for('admin.admin_categories'))

@admin.route('/contact-messages')
@staff_required
def admin_contact_messages():
    delete_form = DeleteForm()
    messages = db.session.execute(db.select(ContactMessage).order_by(ContactMessage.date_posted.desc())).scalars().all()
    return render_template('admin_contact_messages.html', messages=messages, delete_form=delete_form)

@admin.route('/contact-message/edit/<int:message_id>', methods=['GET', 'POST'])
@admin_required
def edit_contact_message(message_id):
    message = db.session.get(ContactMessage, message_id) or abort(404)
    form = ContactMessageEditForm(obj=message)
    if form.validate_on_submit():
        message.name = form.name.data
        message.email = form.email.data
        
        try:
            db.session.commit()
            flash('Message mis à jour avec succès !', 'success')
            return redirect(url_for('admin.admin_contact_messages'))
        except Exception as e:
            db.session.rollback()
            flash(f"Une erreur est survenue lors de la mise à jour : {e}", 'danger')
            return redirect(url_for('admin.edit_contact_message', message_id=message.id))
    return render_template('edit_contact_message.html', message=message, form=form)

@admin.route('/contact-message/delete/<int:message_id>', methods=['POST'])
@admin_required
def delete_contact_message(message_id):
    message_to_delete = db.session.get(ContactMessage, message_id) or abort(404)
    try:
        db.session.delete(message_to_delete)
        db.session.commit()
        flash('Le message a été supprimé avec succès.', 'success')
    except Exception as e:
        db.session.rollback()
        flash("Une erreur est survenue lors de la suppression du message.", 'danger')
    return redirect(url_for('admin.admin_contact_messages'))

@admin.route('/products')
@staff_required
def admin_products():
    delete_form = DeleteForm()
    page = request.args.get('page', 1, type=int)
    q = request.args.get('q', '')
    category_id = request.args.get('category_id', type=int)
    sort_by = request.args.get('sort_by', 'name_asc')
    query = db.select(Product)
    if q:
        query = query.filter(Product.name.ilike(f'%{q}%'))
    if category_id:
        query = query.filter_by(category_id=category_id)
    if sort_by == 'name_asc':
        query = query.order_by(Product.name.asc())
    elif sort_by == 'name_desc':
        query = query.order_by(Product.name.desc())
    elif sort_by == 'price_asc':
        query = query.order_by(Product.price.asc())
    elif sort_by == 'price_desc':
        query = query.order_by(Product.price.desc())
    products = db.paginate(query, page=page, per_page=5)
    categories = Category.query.all()
    return render_template('admin_products.html', products=products, categories=categories, delete_form=delete_form)

@admin.route('/product/add', methods=['GET', 'POST'])
@staff_required
def add_product():
    form = ProductForm()
    form.category.choices = [(c.id, c.name) for c in db.session.execute(db.select(Category).order_by('name')).scalars()]
    if form.validate_on_submit():
        new_product = Product(
            name=form.name.data,
            category_id=form.category.data,
            description=form.description.data,
            price=form.price.data,
            stock=form.stock.data,
            min_stock_threshold=form.min_stock_threshold.data
        )
        images = request.files.getlist(form.image_files.name)
        if images and images[0].filename:
            try:
                # Enregistrer toutes les images comme des objets ProductImage
                for i, image_file in enumerate(images):
                    if image_file and allowed_file(image_file, current_app.config['ALLOWED_EXTENSIONS']):
                        filename = save_image(image_file, current_app.config['UPLOAD_FOLDER'])
                        new_image = ProductImage(image_file=filename, product=new_product, position=i)
                        db.session.add(new_image)
                        # La première image est définie comme principale
                        if i == 0:
                            new_product.image_file = filename
                    elif image_file:
                        flash(f"Le fichier '{image_file.filename}' n'est pas une image valide.", 'danger')
            except Exception as e:
                flash(f"Une erreur est survenue lors de l'enregistrement des images : {e}", 'danger')
                return render_template('add_product.html', form=form)
        db.session.add(new_product)
        db.session.commit()
        flash('Produit ajouté avec succès !', 'success')
        return redirect(url_for('admin.admin_products'))
    return render_template('add_product.html', form=form)

@admin.route('/reply/<int:message_id>', methods=['GET', 'POST'])
@staff_required
def reply_to_contact_message(message_id):
    message = db.session.get(ContactMessage, message_id) or abort(404)
    form = ReplyForm()
    if form.validate_on_submit():
        recipient_email = message.email
        subject = form.subject.data
        message_body = form.message_body.data
        try:
            msg = EmailMessage(subject, message_body, current_app.config['MAIL_DEFAULT_SENDER'], [recipient_email])
            msg.send()
            flash('E-mail envoyé avec succès !', 'success')
            return redirect(url_for('admin.admin_contact_messages'))
        except Exception as e:
            flash(f"Une erreur est survenue lors de l'envoi de l'e-mail : {e}", 'danger')
            return redirect(url_for('admin.reply_to_contact_message', message_id=message.id))
    return render_template('reply_to_contact_message.html', message=message, form=form)

def build_orders_query(filters):
    """Builds the query for orders based on filter and sort parameters."""
    query = db.select(Order).options(db.joinedload(Order.items), db.joinedload(Order.customer))

    status = filters.get('status')
    start_date_str = filters.get('start_date')
    end_date_str = filters.get('end_date')
    sort_by = filters.get('sort_by', 'date_ordered')
    sort_order = filters.get('sort_order', 'desc')

    if status:
        query = query.filter(Order.status == status)
    
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            query = query.filter(Order.date_ordered >= start_date)
        except ValueError:
            flash('Format de date de début invalide. Utilisez AAAA-MM-JJ.', 'danger')

    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            end_date = end_date + timedelta(days=1)
            query = query.filter(Order.date_ordered < end_date)
        except ValueError:
            flash('Format de date de fin invalide. Utilisez AAAA-MM-JJ.', 'danger')

    if hasattr(Order, sort_by):
        if sort_order == 'asc':
            query = query.order_by(getattr(Order, sort_by).asc())
        else:
            query = query.order_by(getattr(Order, sort_by).desc())
    
    return query

@admin.route('/orders')
@staff_required
def admin_orders():
    delete_form = DeleteForm()
    page = request.args.get('page', 1, type=int)
    per_page = 10
    
    filters = {
        'status': request.args.get('status', ''),
        'start_date': request.args.get('start_date', ''),
        'end_date': request.args.get('end_date', ''),
        'sort_by': request.args.get('sort_by', 'date_ordered'),
        'sort_order': request.args.get('sort_order', 'desc')
    }

    query = build_orders_query(filters)
    orders_pagination = db.paginate(query, page=page, per_page=per_page, error_out=False)
    
    statuses = [status[0] for status in db.session.query(Order.status).distinct()]

    return render_template('admin_orders.html', 
                           orders_pagination=orders_pagination, 
                           delete_form=delete_form,
                           statuses=statuses,
                           filters=filters,
                           merge_query_args=merge_query_args)

@admin.route('/order/update_status/<int:order_id>', methods=['POST'])
@staff_required
def update_order_status(order_id):
    order = db.session.get(Order, order_id)
    if not order:
        flash("Commande introuvable.", "danger")
        return redirect(url_for('admin.admin_orders'))
    new_status = request.form.get('status')
    if new_status:
        if order.status_history:
            order.status_history += f";{order.status}"
        else:
            order.status_history = order.status
        order.status = new_status
        db.session.commit()
        try:
            subject = f"Mise à jour du statut de votre commande #{order.id}"
            html_body = render_template('order_status_update.html', order=order, new_status=new_status)
            msg = EmailMessage(subject,
                               html_body,
                               current_app.config['MAIL_DEFAULT_SENDER'],
                               [order.customer.email])
            msg.content_subtype = "html"
            msg.send()
        except Exception as e:
            flash(f"Le statut de la commande a été mis à jour, mais l'envoi de l'e-mail de notification a échoué : {e}", "warning")
        flash(f"Le statut de la commande #{order.id} a été mis à jour.", "success")
    else:
        flash("Veuillez sélectionner un nouveau statut.", "danger")
    return redirect(url_for('admin.admin_orders'))

@admin.route('/users')
@admin_required
def admin_users():
    delete_form = DeleteForm()
    users = db.session.execute(db.select(StaffUser)).scalars().all()
    return render_template('admin_users.html', users=users, delete_form=delete_form)

@admin.route('/staff/add', methods=['GET', 'POST'])
@admin_required
def add_staff():
    form = StaffRegistrationForm()
    if form.validate_on_submit():
        hashed_password = bcrypt.generate_password_hash(form.password.data).decode('utf-8')
        new_staff_user = StaffUser(
            username=form.username.data, 
            email=form.email.data, 
            password=hashed_password, 
            role=form.role.data
        )
        db.session.add(new_staff_user)
        db.session.commit()
        flash('Le membre du personnel a été ajouté avec succès.', 'success')
        return redirect(url_for('admin.admin_users'))
    return render_template('add_staff.html', form=form)

@admin.route('/user/edit/<int:user_id>', methods=['GET', 'POST'])
@admin_required
def edit_user(user_id):
    user = db.session.get(StaffUser, user_id) or abort(404)
    form = StaffUserEditForm(original_username=user.username, original_email=user.email, obj=user)
    if form.validate_on_submit():
        user.username = form.username.data
        user.email = form.email.data
        user.role = form.role.data
        if form.password.data:
            user.password = bcrypt.generate_password_hash(form.password.data).decode('utf-8')
        try:
            db.session.commit()
            flash('Utilisateur du personnel mis à jour avec succès !', 'success')
            return redirect(url_for('admin.admin_users'))
        except Exception as e:
            db.session.rollback()
            flash(f"Une erreur est survenue lors de la mise à jour de l'utilisateur : {e}",'danger')
            return redirect(url_for('admin.edit_user', user_id=user.id))
    return render_template('edit_user.html', form=form, user=user)

@admin.route('/user/delete/<int:user_id>', methods=['POST'])
@admin_required
def delete_user(user_id):
    user_to_delete = db.session.get(StaffUser, user_id) or abort(404)
    if user_to_delete.id == current_user.id:
        flash("Vous ne pouvez pas supprimer votre propre compte !", 'danger')
        return redirect(url_for('admin.admin_users'))
    try:
        db.session.delete(user_to_delete)
        db.session.commit()
        flash('Utilisateur du personnel supprimé avec succès !', 'success')
    except Exception as e:
        db.session.rollback()
        flash("Une erreur est survenue lors de la suppression de l'utilisateur.", 'danger')
    return redirect(url_for('admin.admin_users'))

@admin.route('/customers')
@staff_required
def admin_customers():
    customers = db.session.execute(db.select(Customer)).scalars().all()
    customer_count = len(customers)
    delete_form = DeleteForm()
    return render_template('admin_customers.html', customers=customers, customer_count=customer_count, delete_form=delete_form)

@admin.route('/customer/edit/<int:customer_id>', methods=['GET', 'POST'])
@admin_required
def edit_customer_admin(customer_id):
    customer = db.session.get(Customer, customer_id) or abort(404)
    form = CustomerEditForm(original_username=customer.username, original_email=customer.email, obj=customer)
    if form.validate_on_submit():
        customer.username = form.username.data
        customer.email = form.email.data
        try:
            db.session.commit()
            flash('Client mis à jour avec succès !', 'success')
            return redirect(url_for('admin.admin_customers'))
        except Exception as e:
            db.session.rollback()
            flash(f"Une erreur est survenue lors de la mise à jour du client : {e}", 'danger')
            return render_template('edit_customer_admin.html', form=form, customer=customer)
    return render_template('edit_customer_admin.html', form=form, customer=customer)

@admin.route('/customer/delete/<int:customer_id>', methods=['POST'])
@admin_required
def delete_customer_admin(customer_id):
    customer_to_delete = db.session.get(Customer, customer_id) or abort(404)
    
    if customer_to_delete.orders:
        flash('Vous ne pouvez pas supprimer un client qui a déjà passé des commandes.', 'danger')
        return redirect(url_for('admin.admin_customers'))
        
    try:
        db.session.delete(customer_to_delete)
        db.session.commit()
        flash('Client supprimé avec succès !', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f"Une erreur est survenue lors de la suppression du client : {e}", 'danger')
    return redirect(url_for('admin.admin_customers'))

@admin.route('/product/edit/<int:product_id>', methods=['GET', 'POST'])
@staff_required
def edit_product(product_id):
    product = db.session.get(Product, product_id)
    if not product:
        flash("Produit introuvable.", "danger")
        return redirect(url_for('admin.admin_products'))

    # Correctif pour les produits existants : s'assure que l'image principale a un enregistrement ProductImage.
    # C'est une migration de données "à la volée".
    if product.image_file and product.image_file != 'default.jpg':
        if not any(img.image_file == product.image_file for img in product.images):
            # On décale les positions des images existantes pour faire de la place en première position.
            for img in product.images:
                img.position += 1
            
            # On crée l'enregistrement manquant pour l'image principale.
            main_image_record = ProductImage(
                image_file=product.image_file,
                product_id=product.id,
                position=0  # En première position
            )
            db.session.add(main_image_record)
            db.session.commit()
            
            flash("La structure des images du produit a été mise à jour. Vous pouvez continuer.", "info")
            return redirect(url_for('admin.edit_product', product_id=product_id))

    form = ProductForm()
    form.category.choices = [(c.id, c.name) for c in db.session.execute(db.select(Category).order_by('name')).scalars()]
    if form.validate_on_submit():
        product.name = form.name.data
        product.category_id = form.category.data
        product.description = form.description.data
        product.price = form.price.data
        product.stock = form.stock.data
        product.min_stock_threshold = form.min_stock_threshold.data
        images = request.files.getlist(form.image_files.name)
        if images and images[0].filename:
            try:
                # Déterminer la position de départ pour les nouvelles images
                last_position = db.session.query(db.func.max(ProductImage.position)).filter_by(product_id=product.id).scalar() or -1
                
                for i, image_file in enumerate(images):
                    if image_file and allowed_file(image_file, current_app.config['ALLOWED_EXTENSIONS']):
                        filename = save_image(image_file, current_app.config['UPLOAD_FOLDER'])
                        new_image = ProductImage(image_file=filename, product=product, position=last_position + 1 + i)
                        db.session.add(new_image)
                        
                        # Si le produit n'a pas d'image principale, la première nouvelle image le devient
                        if not product.image_file or product.image_file == 'default.jpg':
                            product.image_file = filename
                            
                    elif image_file:
                        flash(f"Le fichier '{image_file.filename}' n'est pas une image valide.", 'danger')

            except Exception as e:
                flash(f"Une erreur est survenue lors de l'enregistrement des images : {e}", 'danger')
                return render_template('add_product.html', form=form, product=product)
        db.session.commit()
        flash("Produit mis à jour avec succès !", 'success')
        return redirect(url_for('admin.admin_products'))
    elif request.method == 'GET':
        form.name.data = product.name
        form.category.data = product.category_id
        form.description.data = product.description
        form.price.data = product.price
        form.stock.data = product.stock
        form.min_stock_threshold.data = product.min_stock_threshold

    return render_template('add_product.html', form=form, product=product)

@admin.route('/product/<int:product_id>/set-main-image/<int:image_id>', methods=['POST'])
@staff_required
def set_main_image(product_id, image_id):
    product = db.session.get(Product, product_id) or abort(404)
    image = db.session.get(ProductImage, image_id) or abort(404)

    if image not in product.images:
        flash("L'image n'appartient pas à ce produit.", "danger")
        return redirect(url_for('admin.edit_product', product_id=product_id))

    # On définit simplement le nouveau fichier image principal pour le produit.
    product.image_file = image.image_file
    
    db.session.commit()
    flash("L'image principale a été changée avec succès.", "success")
    return redirect(url_for('admin.edit_product', product_id=product_id))


from flask import jsonify, request

@admin.route('/product/<int:product_id>/reorder-images', methods=['POST'])
@staff_required
def reorder_images(product_id):
    product = db.session.get(Product, product_id) or abort(404)
    
    if not request.is_json:
        return jsonify({'success': False, 'message': 'Request must be JSON'}), 400

    new_order = request.json.get('order')
    if not new_order:
        return jsonify({'success': False, 'message': 'Order data missing'}), 400

    try:
        for item in new_order:
            image_id = item.get('id')
            position = item.get('position')
            
            if image_id is None or position is None:
                raise ValueError("Missing image ID or position in order data")

            image = db.session.get(ProductImage, image_id)
            if image and image.product_id == product_id:
                image.position = position
            else:
                raise ValueError(f"Image {image_id} not found or does not belong to product {product_id}")
        
        db.session.commit()
        return jsonify({'success': True, 'message': 'Order updated successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@admin.route('/product/delete_image/<int:image_id>', methods=['POST'])
@admin_required
def delete_image(image_id):
    image_to_delete = db.session.get(ProductImage, image_id) or abort(404)
    product_id = image_to_delete.product_id
    
    # Supprimer de Cloudinary
    if image_to_delete.image_file and 'cloudinary' in image_to_delete.image_file:
        delete_image_from_cloudinary(image_to_delete.image_file)
    try:
        db.session.delete(image_to_delete)
        db.session.commit()
        flash("Image supprimée avec succès !", 'success')
    except Exception as e:
        db.session.rollback()
        flash(f"Une erreur est survenue lors de la suppression de l'image : {e}", 'danger')
    return redirect(url_for('admin.edit_product', product_id=product_id))

@admin.route('/product/delete/<int:product_id>', methods=['POST'])
@admin_required
def delete_product(product_id):
    product_to_delete = db.session.get(Product, product_id) or abort(404)
    # Supprimer toutes les images associées de Cloudinary
    for img in product_to_delete.images:
        delete_image_from_cloudinary(img.image_file)
    try:
        db.session.delete(product_to_delete)
        db.session.commit()
        flash("Produit supprimé avec succès !", 'success')
    except Exception as e:
        db.session.rollback()
        flash("Une erreur est survenue lors de la suppression du produit.", 'danger')
    return redirect(url_for('admin.admin_products'))

@admin.route('/export_contact_messages_excel')
@admin_required
def export_contact_messages_excel():
    messages = db.session.execute(db.select(ContactMessage)).scalars().all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Messages de Contact"
    ws.append(['ID', 'Nom', 'Email', 'Message', "Date d'envoi"])
    for message in messages:
        ws.append([message.id, message.name, message.email, message.message, message.date_posted.strftime('%d/%m/%Y %H:%M')])
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return send_file(excel_file, as_attachment=True, download_name='messages_contact.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@admin.route('/export_orders_excel')
@admin_required
def export_orders_excel():
    orders = db.session.execute(db.select(Order).options(db.joinedload(Order.items).joinedload(OrderItem.product), db.joinedload(Order.customer))).scalars().all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Commandes"
    ws.append(['ID Commande', 'Date', 'Client', 'Email Client', 'Statut', 'ID Produit', 'Nom Produit', 'Quantité', 'Prix Unitaire', 'Total Ligne'])
    for order in orders:
        for item in order.items:
            ws.append([
                order.id,
                order.date_ordered.strftime('%d/%m/%Y %H:%M'),
                order.customer.username,
                order.customer.email,
                order.status,
                item.product.id,
                item.product.name,
                item.quantity,
                item.price_at_purchase,
                item.quantity * item.quantity * item.price_at_purchase
            ])
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return send_file(excel_file, as_attachment=True, download_name='commandes.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@admin.route('/export_products_excel')
@admin_required
def export_products_excel():
    products = db.session.execute(db.select(Product)).scalars().all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Produits"
    ws.append(['ID Produit', 'Nom', 'Catégorie', 'Description', 'Prix', 'Stock', 'Seuil de Stock'])
    for product in products:
        ws.append([
            product.id,
            product.name,
            product.category.name,
            product.description,
            product.price,
            product.stock,
            product.min_stock_threshold
        ])
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return send_file(excel_file, as_attachment=True, download_name='produits.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@admin.route('/export_customers_excel')
@admin_required
def export_customers_excel():
    customers = db.session.execute(db.select(Customer)).scalars().all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Clients Inscrits"
    ws.append(['ID Client', "Nom d'utilisateur", "Email", "Date d'inscription"])
    for customer in customers:
        ws.append([
            customer.id,
            customer.username,
            customer.email,
            customer.date_registered.strftime('%d/%m/%Y %H:%M')
        ])
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return send_file(excel_file, as_attachment=True, download_name='clients_inscrits.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@admin.route('/post/delete/<int:post_id>', methods=['POST'])
@admin_required
def delete_post(post_id):
    post_to_delete = db.session.get(Post, post_id) or abort(404)
    
    if post_to_delete.cover_image and 'cloudinary' in post_to_delete.cover_image:
        delete_image_from_cloudinary(post_to_delete.cover_image)

    try:
        db.session.delete(post_to_delete)
        db.session.commit()
        flash('La réalisation a été supprimée avec succès !', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f"Une erreur est survenue lors de la suppression : {e}", 'danger')
        
    return redirect(url_for('admin.admin_posts'))

@admin.route('/edit-page/<page_name>', methods=['GET', 'POST'])
@admin_required
def edit_page(page_name):
    # Récupère l'objet de contenu de la page, ou le crée s'il n'existe pas.
    content = db.session.get(PageContent, page_name)
    if not content:
        content = PageContent(
            page_name=page_name, 
            title=page_name.replace('_', ' ').title(),
            body='Contenu à remplir.'
        )
        db.session.add(content)
        db.session.commit()

    form = PageContentForm(obj=content)
    if form.validate_on_submit():
        content.title = form.title.data
        content.subtitle = form.subtitle.data
        content.body = form.body.data
        
        if form.image_file.data and isinstance(form.image_file.data, FileStorage):
            try:
                if content.image_file and 'cloudinary' in content.image_file:
                    delete_image_from_cloudinary(content.image_file)
                
                filename = save_image(form.image_file.data, current_app.config['UPLOAD_FOLDER'])
                content.image_file = filename
            except Exception as e:
                flash(f"Erreur lors de l'enregistrement de l'image : {e}", 'danger')
        
        db.session.commit()
        flash('Le contenu de la page a été mis à jour avec succès !', 'success')
        return redirect(url_for('admin.edit_page', page_name=page_name))

    return render_template('admin/edit_page.html', form=form, content=content)


@admin.route('/banners')
@admin_required
def admin_banners():
    banners = db.session.execute(db.select(Banner).order_by(Banner.created_at.desc())).scalars().all()
    delete_form = DeleteForm()
    return render_template('admin/admin_banners.html', banners=banners, delete_form=delete_form)

@admin.route('/banner/add', methods=['GET', 'POST'])
@admin_required
def add_banner():
    form = BannerForm()
    if form.validate_on_submit():
        start_date = None
        if form.start_date.data:
            start_date = datetime.combine(form.start_date.data, datetime.min.time())
        
        end_date = None
        if form.end_date.data:
            end_date = datetime.combine(form.end_date.data, datetime.max.time())

        new_banner = Banner(
            title=form.title.data,
            message=form.message.data,
            link_url=form.link_url.data,
            is_active=form.is_active.data,
            start_date=start_date,
            end_date=end_date,
            position=form.position.data,
            created_at=datetime.now(timezone.utc)
        )
        if form.image.data:
            try:
                filename = save_image(form.image.data, current_app.config['UPLOAD_FOLDER'])
                new_banner.image_file = filename
            except Exception as e:
                flash(f"Une erreur est survenue lors de l'enregistrement de l'image : {e}", 'danger')
                return render_template('admin/add_edit_banner.html', form=form, title="Ajouter une bannière")
        
        db.session.add(new_banner)
        db.session.commit()
        flash('La bannière a été ajoutée avec succès !', 'success')
        return redirect(url_for('admin.admin_banners'))
    return render_template('admin/add_edit_banner.html', form=form, title="Ajouter une bannière")

@admin.route('/banner/edit/<int:banner_id>', methods=['GET', 'POST'])
@admin_required
def edit_banner(banner_id):
    banner = db.session.get(Banner, banner_id) or abort(404)
    form = BannerForm(obj=banner)
    if form.validate_on_submit():
        start_date = None
        if form.start_date.data:
            start_date = datetime.combine(form.start_date.data, datetime.min.time())
        
        end_date = None
        if form.end_date.data:
            end_date = datetime.combine(form.end_date.data, datetime.max.time())

        banner.title = form.title.data
        banner.message = form.message.data
        banner.link_url = form.link_url.data
        banner.is_active = form.is_active.data
        banner.start_date = start_date
        banner.end_date = end_date
        banner.position = form.position.data

        if form.image.data and isinstance(form.image.data, FileStorage):
            try:
                if banner.image_file and 'cloudinary' in banner.image_file:
                    delete_image_from_cloudinary(banner.image_file)
                filename = save_image(form.image.data, current_app.config['UPLOAD_FOLDER'])
                banner.image_file = filename
            except Exception as e:
                flash(f"Une erreur est survenue lors de l'enregistrement de la nouvelle image : {e}", 'danger')
                return render_template('admin/add_edit_banner.html', form=form, title="Modifier une bannière", banner=banner)
        
        db.session.commit()
        flash('La bannière a été mise à jour avec succès !', 'success')
        return redirect(url_for('admin.admin_banners'))
    return render_template('admin/add_edit_banner.html', form=form, title="Modifier une bannière", banner=banner)

@admin.route('/banner/delete/<int:banner_id>', methods=['POST'])
@admin_required
def delete_banner(banner_id):
    banner_to_delete = db.session.get(Banner, banner_id) or abort(404)
    if banner_to_delete.image_file and 'cloudinary' in banner_to_delete.image_file:
        delete_image_from_cloudinary(banner_to_delete.image_file)
    
    try:
        db.session.delete(banner_to_delete)
        db.session.commit()
        flash('La bannière a été supprimée avec succès !', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f"Une erreur est survenue lors de la suppression de la bannière : {e}", 'danger')
    return redirect(url_for('admin.admin_banners'))

@admin.route('/milestones')
@admin_required
def admin_milestones():
    milestones = db.session.execute(db.select(Milestone).order_by(Milestone.order_number.asc())).scalars().all()
    add_form = MilestoneForm()
    delete_form = DeleteForm()
    return render_template('admin/milestones.html', milestones=milestones, add_form=add_form, delete_form=delete_form)

@admin.route('/milestone/add', methods=['POST'])
@admin_required
def add_milestone():
    form = MilestoneForm()
    if form.validate_on_submit():
        new_milestone = Milestone(order_number=form.order_number.data)
        db.session.add(new_milestone)
        db.session.commit()
        flash(f'Le palier {form.order_number.data} a été ajouté avec succès !', 'success')
    else:
        for field, errors in form.errors.items():
            for error in errors:
                flash(f'Erreur pour le champ {getattr(form, field).label.text}: {error}', 'danger')
    return redirect(url_for('admin.admin_milestones'))

@admin.route('/milestone/delete/<int:milestone_id>', methods=['POST'])
@admin_required
def delete_milestone(milestone_id):
    milestone_to_delete = db.session.get(Milestone, milestone_id) or abort(404)
    db.session.delete(milestone_to_delete)
    db.session.commit()
    flash(f'Le palier {milestone_to_delete.order_number} a été supprimé avec succès !', 'success')
    return redirect(url_for('admin.admin_milestones'))

@admin.route('/newsletters')
@admin_required
def newsletters():
    page = request.args.get('page', 1, type=int)
    show_archived = request.args.get('archived', type=bool, default=False)
    
    query = db.select(Newsletter)
    if show_archived:
        query = query.filter_by(archived=True)
    else:
        query = query.filter_by(archived=False)
        
    newsletters = db.paginate(query.order_by(Newsletter.timestamp.desc()), page=page, per_page=10)
    
    send_form = SendForm()
    archive_form = SendForm() # we can reuse the send form for the button
    return render_template('admin/newsletters.html', newsletters=newsletters, send_form=send_form, archive_form=archive_form, show_archived=show_archived)

@admin.route('/newsletter/<int:newsletter_id>/archive', methods=['POST'])
@admin_required
def archive_newsletter(newsletter_id):
    newsletter = db.session.get(Newsletter, newsletter_id) or abort(404)
    newsletter.archived = True
    db.session.commit()
    flash('Newsletter archivée.', 'success')
    return redirect(url_for('admin.newsletters'))

@admin.route('/newsletter/<int:newsletter_id>/unarchive', methods=['POST'])
@admin_required
def unarchive_newsletter(newsletter_id):
    newsletter = db.session.get(Newsletter, newsletter_id) or abort(404)
    newsletter.archived = False
    db.session.commit()
    flash('Newsletter désarchivée.', 'success')
    return redirect(url_for('admin.newsletters', archived=True))

@admin.route('/newsletter/new', methods=['GET', 'POST'])
@admin_required
def new_newsletter():
    form = NewsletterCreationForm()
    if form.validate_on_submit():
        newsletter = Newsletter(subject=form.subject.data, body=form.body.data)
        db.session.add(newsletter)
        db.session.commit()
        flash('Newsletter créée avec succès.', 'success')
        return redirect(url_for('admin.newsletters'))
    return render_template('admin/newsletter_form.html', form=form, title='Créer une newsletter')

@admin.route('/newsletter/<int:newsletter_id>')
@admin_required
def view_newsletter(newsletter_id):
    newsletter = db.session.get(Newsletter, newsletter_id) or abort(404)
    send_form = SendForm()
    return render_template('admin/view_newsletter.html', newsletter=newsletter, send_form=send_form)

@admin.route('/newsletter/<int:newsletter_id>/send', methods=['POST'])
@admin_required
def send_newsletter(newsletter_id):
    newsletter = db.session.get(Newsletter, newsletter_id) or abort(404)
    subscribers = db.session.execute(db.select(NewsletterSubscriber)).scalars().all()
    
    try:
        for subscriber in subscribers:
            msg = EmailMessage(
                subject=newsletter.subject,
                body=newsletter.body,
                from_email=current_app.config['MAIL_DEFAULT_SENDER'],
                to=[subscriber.email],
            )
            msg.send()
        newsletter.sent = True
        db.session.commit()
        flash(f'Newsletter envoyée à {len(subscribers)} abonnés.', 'success')
    except Exception as e:
        flash(f"Erreur lors de l'envoi de la newsletter: {e}", 'danger')

    return redirect(url_for('admin.newsletters'))
